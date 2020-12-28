# Purpose: usign the result of map_item to genereate excel report.
# It will use a definde excel template and fill in with the daily money information

from tony_def import *
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Color, colors, PatternFill
import csv 
import sys
from shutil import copyfile
from map_tab import map_item_struction
import pandas as pd
import numpy as np
import collections
import re
import glob
from map_item_export import *


class gen_rpt_xlsx:
    readable_item_id_start = 1000000000 # unique ID for each item, so far, no certain meaning for this number
    
    def __init__(self):
        self.file_in  = TONY_ALLOUT_DIR+"mapped_item.all.csv"
        self.file_4diff  = "../dat/all_item_backup/"
        self.file_out = "../report/gen_rpt.xlsx"
        self.csv_export = TONY_ALLOUT_DIR+"gen_rpt.db.csv"
        self.template_in = "../src/my_account_template.xlsx"
        self.wb = None
        self.all_df = pd.DataFrame() # raw data from csv
        self.diff_df = pd.DataFrame() # diff data from csv and previuos csv
        self.merge_df = pd.DataFrame() # merge data according to date and type
        self.dayaccount_date_start_position = None # must in 2nd row
        self.dayaccount_date_position = []
        self.dayaccount_type_start_position = None # must in 1st column
        self.dayaccount_type_position = []
        self.readable_tag_num = 1 # default cell number
        self.chk_dayaccount_format_option = [] # option to bypass certain chk
        self.item_export = map_item_export() # export ot andromoney format

    def chk_input_csv(self):

        # check the ext of file
        if (self.file_in[-4:] != ".csv"):
            logging.error("input file type error, yous is %s, expected to be .csv" % self.file_in)

        # check the existence of the file
        try:
            file_in_h = open(self.file_in, newline='')
        except:
            logging.error("An error was found. Either path is incorrect or file doesn't exist!")
        file_in_h.close()

        # if existed, parse to df 
        self.all_df = self.csv2df(csv_in=self.file_in, is_export_db=True) 
    
    # def close_mapped_item_all_csv(self):
    #     self.file_in_h.close()

    def chk_gen_rpt_xlsx(self):
        # chk the template existence
        if (self.template_in[-5:] != ".xlsx"):
            logging.error("input file type error, yous is %s, expected to be .xlsx" % self.template_in)
        
        try:
            wb_template = load_workbook(self.template_in)
        except IOError:
            logging.error("%s is not existed" % self.template_in)
        
        wb_template.close() 
        
        # chk the gen_rpt existence
        if (self.file_out[-5:] != ".xlsx"):
            logging.error("input file type error, yous is %s, expected to be .xlsx" % self.file_out)
        
        # clone the template
        copyfile(src=self.template_in, dst=self.file_out)

        # open the gen_rpt 
        # try:
        #     self.wb = load_workbook(self.file_out)
        # except IOError:
        #     self.wb = Workbook()
        #     logging.info("%s is not existed, we create one" % self.file_out)
        # self.wb.save(self.file_out)
        # self.wb.close()

    def close_gen_rpt_xslx(self):
        self.wb.save(self.file_out)
        self.wb.close()

    def gen_mapped_item_readable_xlsx(self):
        tony_func_proc_disp(msg="Start to gen readable sheet!")
        # always creat a new sheet 
        self.wb = load_workbook(self.file_out)
        sheet_name = "mapped_item_readable"
        if (sheet_name in self.wb.sheetnames):
            sheet = self.wb[sheet_name]
            self.wb.remove_sheet(sheet)
            logging.info("Remove the %s and create a new one!" % sheet_name)
        self.wb.create_sheet(sheet_name)
        sheet = self.wb[sheet_name]
        
        # to mark the duplicated item based on date type adn expense
        # for idx in self.all_df[self.all_df.loc[:,['date','type','expense']].astype(str).duplicated()].index:
        for idx in self.all_df[self.all_df.astype(str).duplicated()].index:
            self.all_df.loc[idx,'status'] = 'Duplicated'
            logging.error("This item is duplicated!! %s" % self.all_df.loc[idx,:])

        # parse csv to readable excel
        # write content
        map_tab = map_item_struction()
        row=1
        col=0
        cur_date = None
        nxt_date = None
        colorFill0 = PatternFill(start_color="00D0D0D0", end_color="00D0D0D0", fill_type="solid")
        colorFill1 = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
        colorFill  = [colorFill0,colorFill1]
        colorFill_error = PatternFill(start_color="00FF0000", end_color="00FF0000", fill_type="solid")
        for i in range(self.all_df.shape[0]):
            line = self.all_df.iloc[i,:]
            is_duplicated_item = self.all_df.loc[i,'status'] == 'Duplicated'
            #to swap the cell color for adjancnet line
            nxt_date = line['date']
            if (nxt_date != cur_date):
                cur_date = nxt_date
                colorFill = colorFill[::-1]

            # to expand the tags and support different tags item 
            line_val = list(line.values)
            tmp = line_val.pop(map_tab.item_struc_name.index('tag'))
            for ii in range (0,self.readable_tag_num):
                if (ii<len(tmp)):
                    line_val.insert(map_tab.item_struc_name.index('tag'),tmp[len(tmp)-ii-1])
                else:
                    line_val.insert(map_tab.item_struc_name.index('tag')+len(tmp),"")

            #write out
            col=0
            item_id = 'ID_' + str(self.readable_item_id_start) # start from 1000000000
            self.readable_item_id_start += 1
            sheet.cell(row = (row+1), column=(col+1), value=item_id) # sheet start from 1
            sheet.cell(row = (row+1), column=(col+1)).fill = colorFill[0] # to distinguish the daliy item 
            for j in line_val:
                sheet.cell(row = (row+1), column=(col+2), value=str(j)) # sheet start from 1
                sheet.cell(row = (row+1), column=(col+2)).fill = colorFill[0] if not (is_duplicated_item) else colorFill_error # to distinguish the daliy item 
                col+=1
                # print (j)
            row+=1
        
        # write classify
        row=0
        col=0
        map_tab = map_item_struction()
        colorFill = PatternFill(start_color="AACF91", end_color="AACF91", fill_type="solid")
        sheet.cell(row = 1, column= 1, value="ITEM_ID") # add item_id in readable sheet
        sheet.cell(row = 1, column= 1).fill = colorFill 
        for i in map_tab.item_struc_name:
            if (i == 'note'):
                sheet.cell(row = (row+1), column=(col+2+self.readable_tag_num-1), value=i) # sheet start from 1
                sheet.cell(row = (row+1), column=(col+2+self.readable_tag_num-1)).fill = colorFill
            else:
                sheet.cell(row = (row+1), column=(col+2), value=i) # sheet start from 1
                sheet.cell(row = (row+1), column=(col+2)).fill = colorFill
            col += 1
        self.close_gen_rpt_xslx()

    def gen_mapped_item_diff_xlsx(self,date4diff="PREVIOUS_ALL_ITEM"):
        """
        if date4diff is not given to a certain date (date4diff="yyyy_mm_dd"), it will pick up the previous one for diff
        """
        tony_func_proc_disp(msg="Start to gen diff sheet!")
        
        # gen a df for date4diff
        self.wb = load_workbook(self.file_out)
        if date4diff == "PREVIOUS_ALL_ITEM":
            file_in_path = self.file_4diff+"*.csv"
            file_in_all = glob.glob(file_in_path)
            file_in_all.sort() # always inplace
            logging.debug("All of mapped_all_item is %s" % file_in_all)
            if (len(file_in_all) < 2):
                logging.error("No previous all_item is available!!")
                raise TypeError("Error")
            self.file_4diff = file_in_all[-2] # -1 is the all_item of current date
        else:
            self.file_4diff += "mapped_item_" + date4diff + ".all.csv"
        logging.debug("File for diff is %s" % self.file_4diff)
        diff_df = self.csv2df(csv_in=self.file_4diff)

        # diff two df
        tmp_pd = pd.concat([self.all_df,diff_df])
        # tmp_pd.sort_values(by=['date','type','expense'],ascending=[True,True,False],inplace=True)
        logging.debug ("the tmp_pd               %s" % tmp_pd)
        logging.debug ("the index not duplicated %s" % tmp_pd.astype(str).drop_duplicates(keep=False).index)
        logging.debug ("the index duplicated index %s" % tmp_pd.astype(str).duplicated(keep=False))
        self.diff_df = tmp_pd.iloc[tmp_pd.astype(str).drop_duplicates(keep=False).index] # change type to str -> drop duplcate -> rtn the index 
        logging.debug ("the diff_df                %s" % self.diff_df                   )

        # always creat a new sheet 
        sheet_name = "mapped_item_diff"
        if (sheet_name in self.wb.sheetnames):
            sheet = self.wb[sheet_name]
            self.wb.remove_sheet(sheet)
            logging.info("Remove the %s and create a new one!" % sheet_name)
        self.wb.create_sheet(sheet_name)
        sheet = self.wb[sheet_name]

        # parse csv to diff excel
        # write content
        map_tab = map_item_struction()
        row=1
        col=0
        cur_date = None
        nxt_date = None
        colorFill0 = PatternFill(start_color="00D0D0D0", end_color="00D0D0D0", fill_type="solid")
        colorFill1 = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
        colorFill  = [colorFill0,colorFill1]
        for i in range(self.diff_df.shape[0]):
            line = self.diff_df.iloc[i,:]
            #to swap the cell color for adjancnet line
            nxt_date = line['date']
            if (nxt_date != cur_date):
                cur_date = nxt_date
                colorFill = colorFill[::-1]

            # to expand the tags and support different tags item 
            line_val = list(line.values)
            tmp = line_val.pop(map_tab.item_struc_name.index('tag'))
            for ii in range (0,self.readable_tag_num):
                if (ii<len(tmp)):
                    line_val.insert(map_tab.item_struc_name.index('tag'),tmp[len(tmp)-ii-1])
                else:
                    line_val.insert(map_tab.item_struc_name.index('tag')+len(tmp),"")

            #write out
            col=0
            item_id = str(row)
            sheet.cell(row = (row+1), column=(col+1), value=item_id) # sheet start from 1
            sheet.cell(row = (row+1), column=(col+1)).fill = colorFill[0] # to distinguish the daliy item 
            for j in line_val:
                sheet.cell(row = (row+1), column=(col+2), value=str(j)) # sheet start from 1
                sheet.cell(row = (row+1), column=(col+2)).fill = colorFill[0] # to distinguish the daliy item 
                col+=1
                # print (j)
            row+=1
        
        # write classify
        row=0
        col=0
        map_tab = map_item_struction()
        colorFill = PatternFill(start_color="AACF91", end_color="AACF91", fill_type="solid")
        sheet.cell(row = 1, column= 1, value="ITEM_ID") # add item_id in readable sheet
        sheet.cell(row = 1, column= 1).fill = colorFill 
        for i in map_tab.item_struc_name:
            if (i == 'note'):
                sheet.cell(row = (row+1), column=(col+2+self.readable_tag_num-1), value=i) # sheet start from 1
                sheet.cell(row = (row+1), column=(col+2+self.readable_tag_num-1)).fill = colorFill
            else:
                sheet.cell(row = (row+1), column=(col+2), value=i) # sheet start from 1
                sheet.cell(row = (row+1), column=(col+2)).fill = colorFill
            col += 1
        
        self.close_gen_rpt_xslx()
    
    
    def chk_dayaccount_format(self,sheet,date_row='2'):
        """
        to extract the type position in the template sheet
        chk the extract type is as same as item_type
        """
        # extract type position
        sh_range = sheet['A'] # read the 1st column
        is_1st_cell = False
        is_last_cell = False
        for cell in sh_range:
            # print (cell)
            # print (cell.value)
            # print (cell.coordinate)
            if (cell.coordinate != 'A1'): # the 1st row cannot be type
                if ((cell.value != None) and (self.dayaccount_type_start_position == None)):
                    is_1st_cell = True
                    self.dayaccount_type_start_position = cell.coordinate
                    self.dayaccount_type_position.append(cell.value)
                elif (is_last_cell):
                    if (cell.value == 'Summary'): # list some exception in A cloumn
                        pass
                    elif (cell.value != None):
                        logging.error("The type should be None @ %s" % cell.coordinate)
                        raise TypeError("ERROR!!")
                elif (is_1st_cell):
                    if (cell.value == None):
                        if (is_last_cell == False):
                            is_last_cell = True
                        else:
                            logging.error("The type should be continuous!")
                            raise TypeError("ERROR!!")
                    else:
                        self.dayaccount_type_position.append(cell.value)
        
        # extract date position
        sh_range = sheet[date_row] # read the 2nd row
        is_1st_cell = False
        is_last_cell = False
        for cell in sh_range:
            # print (cell)
            # print (cell.value)
            # print (cell.coordinate)
            if (cell.coordinate != ('A'+date_row)): # the 1st column cannot be date
                if ((cell.value != None) and (self.dayaccount_date_start_position == None)):
                    is_1st_cell = True
                    self.dayaccount_date_start_position = cell.coordinate
                # elif (is_last_cell):
                #     if (cell.value == 'Any Exception?'): # list some exception in 2 row
                #         pass
                #     elif (cell.value != None):
                #         logging.error("The type should be None @ %s" % cell.coordinate)
                #         raise TypeError("ERROR!!")
                if (is_1st_cell):
                    if (cell.value == None):
                        self.dayaccount_date_position.append('None') # fill in reserved word
                    else:
                        split_pattern = r'\W+|_'
                        tmp = re.split(split_pattern,str(cell.value))
                        # print (tmp)
                        date = tmp[0] + "_" + tmp[1] + "_" + tmp[2] 
                        chk_pattern = r'\d{4}_\d{2}_\d{2}'
                        if not ( re.match(chk_pattern,date,0) ):
                            logging.error("date formate should be yyyy_mm_dd, and yours is %s" % date)
                            raise TypeError("Error")
                        self.dayaccount_date_position.append(date)

        # check template
        # existence
        if (len(self.dayaccount_type_position) == 0):
            raise TypeError(" %s is not existed" % self.dayaccount_type_position)
        if (len(self.dayaccount_date_position) == 0):
            raise TypeError(" %s is not existed" % self.dayaccount_date_position)

        # sync to map_tab, type in excel cannot exceed item_type 
        map_tab = map_item_struction()
        A = set(self.dayaccount_type_position)
        B = set(map_tab.item_type)
        A_diff_B = A.difference(B)
        B_diff_A = B.difference(A)
        if (len(A_diff_B)!=0):
            logging.error("A: %s" % A)
            logging.error("B: %s" % B)
            logging.error("A Diff B: %s" % A.difference(B))
            logging.error("B Diff A: %s" % B.difference(A))
            raise TypeError(" type content is not the same, please sync it!!")
        if (len(B_diff_A)!=0):
            logging.info("A: %s" % A)
            logging.info("B: %s" % B)
            logging.info("A Diff B: %s" % A.difference(B))
            logging.info("B Diff A: %s" % B.difference(A))
        # if (set(self.dayaccount_type_position) != set(map_tab.item_type)):
        #     A = set(self.dayaccount_type_position)
        #     B = set(map_tab.item_type)
        #     logging.error("A: %s" % A)
        #     logging.error("B: %s" % B)
        #     logging.error("A Diff B: %s" % A.difference(B))
        #     logging.error("B Diff A: %s" % B.difference(A))
        #     raise TypeError(" type content is not the same, please sync it!!")
        
        # no duplicate
        duplicate = [item for item, count in collections.Counter(self.dayaccount_type_position).items() if count > 1]
        if (len(duplicate) != 0):
            logging.error("dulicate item are %s: " % duplicate)
            raise TypeError(" type content is duplcated!!")
        duplicate = [item for item, count in collections.Counter(self.dayaccount_date_position).items() if ((item != 'None') and (count > 1))]
        if (len(duplicate) != 0):
            logging.error("dulicate item are %s: " % duplicate)
            raise TypeError(" type content is duplcated!!")

        # no invalide content
        if ('IGNORE_CHK_INVALID_CONTENT' in self.chk_dayaccount_format_option):
            logging.info("IGNORE_CHK_INVALID_CONTENT")
            pass
        else:
            # pattern = re.compile(r'\d{4}')
            # tmp_list = map((lambda x: pattern.search(x,0)),self.dayaccount_date_position) # extract the year part
            # year_list = [i.group(0) for i in tmp_list if (i!=None)] # extract the year part
            # year_cnt = collections.Counter(year_list) # cal the days in each year
            # # print (year_cnt)
            # # print (year_cnt.keys())
            # # print (year_cnt.values())
            # for k,v in year_cnt.items(): # because there are no duplicated date, the total number of year is fixed
            #     if ( (int(k)%4 == 0) and (v != 366) ):
            #         logging.error("for year: %s, the total days should be 366" % k)
            #         raise TypeError("Data error")
            #     elif ( (int(k)%4 == 1) and (v != 365) ):
            #         logging.error("for year: %s, the total days should be 365" % k)
            #         raise TypeError("Data error")
            pass # no need to check this 
        # for i in year_set:
        #     if ((int(i)%4 == 0) and (year_cnt.values() != 366)): # have ...._02_29
        #         year_cnt.values()
        # for i in year_list:
        #     print (i)
        
        # invalid_list = ['02_30','02_31','04_31','06_31','09_31','11_31']
        # split_pattern = r'_'
        # chk_list = [re.split(split_pattern,item) for item in invalid_list]
        # # search_pattern = r'\d{2}_\d{2}'
        # # have to change to compile and set the start position of the re.search
        # pattern = re.compile(r'\d{2}_\d{2}')
        # # haha = filter((lambda x: pattern.search(str(x[5:]))),self.dayaccount_date_position)
        # data_list = map((lambda x: pattern.search(x,5)),self.dayaccount_date_position)
        # haha = filter( (lambda x: (x in chk_list)),data_list)
        # print (haha)
        # for i in haha:
        #     print (i)
        # src_list = [pattern.search(item, 5) for item in self.dayaccount_date_position if item != None]
        # print (chk_list)
        # print (src_list)
        # date = tmp[0] + "_" + tmp[1] + "_" + tmp[2] 
        # chk_pattern = r'\d{4}_\d{2}_\d{2}'
        # if not ( re.match(chk_pattern,date,5) ):
        #     logging.error("date formate should be yyyy_mm_dd, and yours is %s" % date)
        #     raise TypeError("Error")


    def locate_dayaccount_date(self,sheet,date,ofs_only=1):
        """
        return cell locate of the date or the offset relative to baseline
        """
        ofs_col = self.dayaccount_date_position.index(date)
        cell = sheet[self.dayaccount_date_start_position].offset(row=0,column=ofs_col)
        logging.debug (cell)
        return ( cell.column_letter if (ofs_only) else cell.coordinate )
    
    def locate_dayaccount_type(self,sheet,type,ofs_only=1):
        """
        return cell locate of the type or the offset relative to baseline
        """
        ofs_row = self.dayaccount_type_position.index(type)
        cell = sheet[self.dayaccount_type_start_position].offset(row=ofs_row,column=0)
        logging.debug (cell)
        return ( cell.row if (ofs_only) else cell.coordinate )

    def locate_dayaccount_link(self,sheet,date,type):
        """
        return cell link dst in gen_rpt.xlsx#mapped_item_readable
        """
        link_sheet_name = "mapped_item_readable"
        if (link_sheet_name not in self.wb.sheetnames):
            logging.error(" sheet name: %s cannot found" % link_sheet_name)
            raise TypeError("Error")
        match_date_df = self.all_df.loc[self.all_df['date'] == date]
        match_type_df = match_date_df.loc[self.all_df['type'] == type]
        # match_date.sort_index(inplace=True) # warn!?
        link_ofs = min(match_type_df.index.values) + 2 # content start from row 2
        logging.debug("link to %s: %s -> \n%s" % (link_sheet_name,link_ofs,match_type_df.loc[:,['name','expense']]))
        link = "gen_rpt.xlsx#" + link_sheet_name + "!" + "A" + str(link_ofs)
        return link
        # return match_date.index[0]
        # print (match_date.iloc[0,:].index)
        # print ("000",[self.all_df['date'] == '2020_07_05']) # list the match series
        # print ("111",self.all_df.loc[self.all_df['date'] == '2020_07_05']) # filter out those row in True
        # print ("222",aaa.iloc[[0]]) 
        # print (self.all_df.loc[self.all_df['date'] == '2020_07_05'])
        # pass

    # def gen_dayaccount_xlsx(self):
    #     tony_func_proc_disp(msg="Start to gen dayaccount sheet!")
    #     # make sure the template is there 
    #     self.wb = load_workbook(self.file_out)
    #     sheet_name = "template"
    #     if (sheet_name not in self.wb.sheetnames):
    #         logging.error(" %s is not existed" % sheet_name)
    #         raise TypeError(" %s is not existed" % sheet_name)
    #     sh_template = self.wb[sheet_name]

    #     # always clone from template to a new sheet in gen_rpt
    #     sheet_name = "dayaccount"
    #     if (sheet_name in self.wb.sheetnames):
    #         sheet = self.wb[sheet_name]
    #         self.wb.remove_sheet(sheet)
    #         logging.info("Remove the %s and create a new one!" % sheet_name)
    #     sh_dayaccount = self.wb.copy_worksheet(sh_template)
    #     sh_dayaccount.title = sheet_name
    #     self.chk_dayaccount_format(sh_dayaccount)
        
    #     logging.debug("sh_dayaccount.title %s" % sh_dayaccount.title)
    #     logging.debug("sh_dayaccount.max_row %s" % sh_dayaccount.max_row)
    #     logging.debug("sh_dayaccount.max_column %s" % sh_dayaccount.max_column)

    #     # to merger the expense of the same date and type
    #     df_column = ['date','type','expense']
    #     cur_pair = ()
    #     for i in range(self.all_df.shape[0]+1): # +1 so that the i can reach the end fo col
    #         if (i==self.all_df.shape[0]):
    #             sum_exp  = np.sum(df_merge.loc[:,['expense']])
    #             df_tmp = pd.DataFrame([list(cur_pair)+[sum_exp[0]]],columns=df_column)
    #             df = df.append(df_tmp)
    #         else:
    #             line = self.all_df.iloc[i,:]
    #             df_nxt = pd.DataFrame([[line.date,line.type,line.expense]],columns=df_column)
    #             nxt_pair = (line.date,line.type)
    #             if (i==0):
    #                 df = pd.DataFrame()
    #                 df_merge = pd.DataFrame()
    #                 df_merge = df_merge.append(df_nxt)
    #                 cur_pair  = (line.date,line.type)
    #             # elif (i==self.all_df.shape[0]-1):
    #             #     df_merge = df_merge.append(df_nxt)
    #             #     sum_exp  = np.sum(df_merge.loc[:,['expense']])
    #             #     df_tmp = pd.DataFrame([list(cur_pair)+[sum_exp[0]]],columns=df_column)
    #             #     df = df.append(df_tmp)
    #             elif (nxt_pair != cur_pair):
    #                 sum_exp  = np.sum(df_merge.loc[:,['expense']])
    #                 df_tmp = pd.DataFrame([list(cur_pair)+[sum_exp[0]]],columns=df_column)
    #                 df = df.append(df_tmp)
    #                 df_merge = df_nxt
    #                 cur_pair  = (line.date,line.type)
    #             else:
    #                 df_merge = df_merge.append(df_nxt)

    #         logging.debug ("df_nxt %s" % df_nxt)
    #         logging.debug ("df_merge %s" % df_merge)
    #     self.merge_df = df
    #     logging.debug ("df is %s" % df)

    #     # tranlate from df to xsl

    #     # write to dayaccount sheet
    #     tony_func_proc_disp(msg="Write out to dayccount sheet!")
    #     for i in range(self.merge_df.shape[0]):
    #         line = self.merge_df.iloc[i,:]
    #         logging.info ("Write out to dayaccount @ %s" % line['date'])
    #         row = self.locate_dayaccount_date(sheet=sh_dayaccount,date=line['date']) # determine row
    #         col = self.locate_dayaccount_type(sheet=sh_dayaccount,type=line['type']) # determine col
    #         pos = row + str(col)
    #         link = self.locate_dayaccount_link(sheet=sh_dayaccount,date=line['date'],type=line['type'])
    #         sh_dayaccount[pos].value=line['expense']
    #         sh_dayaccount[pos].hyperlink=link
    #         sh_dayaccount[pos].font = Font(u='single', color=colors.BLUE)
        
    #     self.close_gen_rpt_xslx()
    
    def gen_dayaccount_xlsx(self):
        tony_func_proc_disp(msg="Start to gen dayaccount sheet!")

        map_tab = map_item_struction()
        
        # extract the yyyy_mm_dd
        date_df = self.all_df.loc[:,'date'].copy()
        # date_wi_dd = [ date_df[i][:-3] for i in range(len(date_df)) ] # how to manipulate each item in Series?? 
        # date_wi_dd_df = pd.Series(date_wi_dd)
        day_expense = pd.Series([ i*0  for i in range(0,len(map_tab.item_type))],index=map_tab.item_type)
        day_expense_df = pd.DataFrame()
        day_df = date_df.drop_duplicates(keep='first')
        for i, item in zip (range(len(day_df)), day_df):
            day_expense_df.insert(loc=i,column=item,value=day_expense,allow_duplicates=True)

        # cal 
        for i in range(len(self.all_df)):
            logging.debug(i)
            logging.debug(date_df[i])
            day_expense_df.loc[self.all_df['type'][i],date_df[i]] += self.all_df['expense'][i]
            logging.debug(day_expense_df.loc[self.all_df['type'][i],date_df[i]])
        
        # write out 
        with pd.ExcelWriter(self.file_out, engine='openpyxl') as writer: # the way not to overwrite the existed excel
            writer.book = load_workbook(self.file_out)
            day_expense_df.to_excel(writer, "day_expense")
    
    
    def gen_monthaccount_xlsx(self):
        tony_func_proc_disp(msg="Start to gen monthaccount sheet!")

        map_tab = map_item_struction()
        
        # extract the yyyy_mm
        date_df = self.all_df.loc[:,'date'].copy()
        date_wo_dd = [ date_df[i][:-3] for i in range(len(date_df)) ] # how to manipulate each item in Series?? 
        date_wo_dd_df = pd.Series(date_wo_dd)
        mon_expense = pd.Series([ i*0  for i in range(0,len(map_tab.item_type))],index=map_tab.item_type)
        mon_expense_df = pd.DataFrame()
        mon_df = date_wo_dd_df.drop_duplicates(keep='first')
        for i, item in zip (range(len(mon_df)), mon_df):
            mon_expense_df.insert(loc=i,column=item,value=mon_expense,allow_duplicates=True)

        # cal 
        for i in range(len(self.all_df)):
            logging.debug(i)
            logging.debug(date_wo_dd_df[i])
            mon_expense_df.loc[self.all_df['type'][i],date_wo_dd_df[i]] += self.all_df['expense'][i]
            logging.debug(mon_expense_df.loc[self.all_df['type'][i],date_wo_dd_df[i]])
        
        # write out 
        with pd.ExcelWriter(self.file_out, engine='openpyxl') as writer: # the way not to overwrite the existed excel
            writer.book = load_workbook(self.file_out)
            mon_expense_df.to_excel(writer, "mon_expense")
    
    
    def gen_yearaccount_xlsx(self):
        tony_func_proc_disp(msg="Start to gen yearaccount sheet!")

        map_tab = map_item_struction()
        
        # extract the yyyy
        date_df = self.all_df.loc[:,'date'].copy()
        date_wo_mmdd = [ date_df[i][:-6] for i in range(len(date_df)) ] # how to manipulate each item in Series?? 
        date_wo_mmdd_df = pd.Series(date_wo_mmdd)
        year_expense = pd.Series([ i*0  for i in range(0,len(map_tab.item_type))],index=map_tab.item_type)
        year_expense_df = pd.DataFrame()
        year_df = date_wo_mmdd_df.drop_duplicates(keep='first')
        for i, item in zip (range(len(year_df)), year_df):
            year_expense_df.insert(loc=i,column=item,value=year_expense,allow_duplicates=True)

        # cal 
        for i in range(len(self.all_df)):
            logging.debug(i)
            logging.debug(date_wo_mmdd_df[i])
            year_expense_df.loc[self.all_df['type'][i],date_wo_mmdd_df[i]] += self.all_df['expense'][i]
            logging.debug(year_expense_df.loc[self.all_df['type'][i],date_wo_mmdd_df[i]])
        
        # write out 
        with pd.ExcelWriter(self.file_out, engine='openpyxl') as writer: # the way not to overwrite the existed excel
            writer.book = load_workbook(self.file_out)
            year_expense_df.to_excel(writer, "year_expense")

    
    # def gen_totalaccount_xlsx(self):
    #     tony_func_proc_disp(msg="Start to gen totalaccount sheet!")

    #     map_tab = map_item_struction()
        
    #     total_expense = pd.Series([ i*0  for i in range(0,len(map_tab.item_type))],index=map_tab.item_type)
    #     total_expense_df = pd.DataFrame()
    #     total_expense_df.insert(loc=0,column=0,value=total_expense,allow_duplicates=True)

    #     # cal 
    #     for i in range(len(self.all_df)):
    #         logging.debug(i)
    #         total_expense_df.loc[self.all_df['type'][i],0] += self.all_df['expense'][i]
        
    #     # write out 
    #     with pd.ExcelWriter(self.file_out, engine='openpyxl') as writer: # the way not to overwrite the existed excel
    #         writer.book = load_workbook(self.file_out)
    #         total_expense_df.to_excel(writer, "total_expense")
    
    def gen_totalaccount_xlsx(self):
        tony_func_proc_disp(msg="Start to gen totalaccount sheet!")

        map_tab = map_item_struction()
        
        # extract the yyyy_mm
        date_df = self.all_df.loc[:,'date'].copy()
        date_wo_dd = [ date_df[i][:-3] for i in range(len(date_df)) ] # how to manipulate each item in Series?? 
        date_wo_dd_df = pd.Series(date_wo_dd)
        mon_expense = pd.Series([ i*0  for i in range(0,len(map_tab.item_type))],index=map_tab.item_type)
        mon_expense_df = pd.DataFrame()
        mon_df = date_wo_dd_df.drop_duplicates(keep='first')
        for i, item in zip (range(len(mon_df)), mon_df):
            mon_expense_df.insert(loc=i,column=item,value=mon_expense,allow_duplicates=True)
        
        # extract the yyyy
        date_wo_mmdd = [ date_df[i][:-6] for i in range(len(date_df)) ] # how to manipulate each item in Series?? 
        date_wo_mmdd_df = pd.Series(date_wo_mmdd)
        year_expense = pd.Series([ i*0  for i in range(0,len(map_tab.item_type))],index=map_tab.item_type)
        year_expense_df = pd.DataFrame()
        year_df = date_wo_mmdd_df.drop_duplicates(keep='first')
        for i, item in zip (range(len(year_df)), year_df):
            year_expense_df.insert(loc=i,column=item,value=year_expense,allow_duplicates=True)
    
        total_expense = pd.Series([ i*0  for i in range(0,len(map_tab.item_type))],index=map_tab.item_type)
        total_expense_df = pd.DataFrame()
        total_expense_df.insert(loc=0,column=0,value=total_expense,allow_duplicates=True)

        # cal 
        for i in range(len(self.all_df)):
            logging.debug(i)
            logging.debug(date_wo_dd_df[i])
            mon_expense_df.loc[self.all_df['type'][i],date_wo_dd_df[i]] += self.all_df['expense'][i]
            year_expense_df.loc[self.all_df['type'][i],date_wo_mmdd_df[i]] += self.all_df['expense'][i]
            total_expense_df.loc[self.all_df['type'][i],0] += self.all_df['expense'][i]
            logging.debug(mon_expense_df.loc[self.all_df['type'][i],date_wo_dd_df[i]])
        
        # write out 
        with pd.ExcelWriter(self.file_out, engine='openpyxl') as writer: # the way not to overwrite the existed excel
            writer.book = load_workbook(self.file_out)
            mon_expense_df.to_excel(writer, "mon_expense")
            year_expense_df.to_excel(writer, "year_expense")
            total_expense_df.to_excel(writer, "total_expense")
    
    def gen_item_export(self):
        tony_func_proc_disp(msg=" Start to export to andromoney!")
        self.item_export.do_item_export()
        tony_func_proc_disp(msg=" Done export!")

    
    def csv2df (self,csv_in, is_export_db=False):
        # parse csv to DataFrame
        tony_func_proc_disp(msg="Transfer csv to df!")
        df_out = pd.DataFrame() # raw data from csv
        map_tab = map_item_struction()
        map_item = map_tab.gen_empty_item()
        is_classify = True
        df_column = []
        with open(csv_in, mode='r', newline='', encoding='UTF-8-sig') as csvfile:
            reader = csv.reader(csvfile)
            row=0
            for line in reader:
                chk_pattern = map_tab.map_tab_cmt_prefix 
                if ( re.match(chk_pattern,line[0],0) ):
                    tmp = ( str(w)+',' for w in line)
                    cmt = ''.join(tmp)
                    logging.info("cmt: %s" % cmt)
                else:
                    if (is_classify == True):
                        is_classify = False
                        df_column = list(line)
                        df_out = pd.DataFrame(columns=df_column)
                    else:
                        map_item = self.csv2mappeditem(line)
                        # print (df_column)
                        # print (map_item)
                        # sub = pd.DataFrame(map_item,columns=df_column,index=[row])
                        sub = pd.DataFrame([map_item],columns=df_column,index=[row-1]) # row start from 0
                        df_out = df_out.append(sub)
                        # print (sub)
                        # print (df)
                        row+=1
        if (not df_out.empty):
            logging.debug("==============%s=====before_sort===============" % df_out)
            logging.debug(df_out.loc[:,['name','date','type','expense','tag']])
            logging.debug("===============================================")
            df_out.sort_values(by=['date','type','expense'],ascending=[True,True,False],inplace=True)
            df_out.reset_index(drop=True,inplace=True) # must be inpalced and drop old one
            logging.debug("==============%s===f=after_sort================" % df_out)
            logging.debug(df_out.loc[:,['name','date','type','expense','tag']])
            logging.debug("===============================================")

        # export to csv
        if (is_export_db):
            access = 'w' 
            with open(self.csv_export, access, newline='', encoding='UTF-8-sig') as csvfile:
                logging.debug ("touch the file only")
            df_out.to_csv(self.csv_export,index=0)

        # try to import 
        # df_import = pd.read_csv(self.csv_export,index_col=False)

        return df_out

    
    def csv2mappeditem (self,csvitem):
        map_tab = map_item_struction()
        map_item = map_tab.gen_empty_item()
        map_item[map_tab.item_struc_name.index("name")]      = csvitem[0]
        map_item[map_tab.item_struc_name.index("date")]      = csvitem[1] 
        map_item[map_tab.item_struc_name.index("category")]  = csvitem[2] 
        map_item[map_tab.item_struc_name.index("type")]      = csvitem[3] 
        map_item[map_tab.item_struc_name.index("expense")]   = csvitem[4] 
        map_item[map_tab.item_struc_name.index("income")]    = csvitem[5] 
        map_item[map_tab.item_struc_name.index("source")]    = csvitem[6] 
        map_item[map_tab.item_struc_name.index("status")]    = csvitem[7] 
        map_item[map_tab.item_struc_name.index("location")]  = csvitem[8] 
        map_item[map_tab.item_struc_name.index("tag")]       = eval(csvitem[9])
        map_item[map_tab.item_struc_name.index("note")]      = eval(csvitem[10])
        map_tab.chg_mapped_item_type(map_item)

        # to support multi tags in readable sheet
        tag_num = len(map_item[map_tab.item_struc_name.index("tag")])
        self.readable_tag_num = tag_num if (tag_num > self.readable_tag_num) else self.readable_tag_num

        return map_item.copy()
    



            
        

if __name__ == "__main__":
    tony_func_proc_disp(msg=" Start to test gen_rpt only!!")
    
    test = gen_rpt_xlsx()
    test.chk_input_csv()
    test.gen_item_export()
    test.chk_gen_rpt_xlsx()
    test.gen_mapped_item_readable_xlsx()
    test.gen_mapped_item_diff_xlsx()
    test.gen_dayaccount_xlsx()
    test.gen_totalaccount_xlsx()
    
    tony_func_proc_disp(msg=" Done test gen_rpt only!!")
